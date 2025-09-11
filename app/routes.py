from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, current_app, make_response
from flask_login import login_required, current_user
from werkzeug.security import check_password_hash
import psycopg2
from sqlalchemy import text
import os
import time

from .models import User, Role, WalletTransaction
from .wallet import ensure_qr_for_user, change_balance, generate_qr_data_uri
from . import db, cache
from .audit import log_event
from .email_utils import send_credentials_email

main_bp = Blueprint("main", __name__)


@main_bp.get("/")
def index():
	if current_user.is_authenticated:
		return redirect(url_for("main.dashboard"))
	return render_template("index.html")


@main_bp.get("/mobile")
def mobile_access():
	"""Mobile access page with QR code and instructions"""
	return render_template("mobile_access.html")




@main_bp.get("/debug/users")
def debug_users():
	users = User.query.all()
	result = []
	for user in users:
		result.append({
			"id": user.id,
			"username": user.username,
			"role": user.role.value if user.role else None,
			"email": user.email,
			"balance": user.balance
		})
	return {"users": result, "count": len(result)}


@main_bp.post("/admin/delete-user")
@login_required
def admin_delete_user():
	"""Delete a user from the system"""
	if current_user.role != Role.ADMIN:
		flash("Unauthorized", "danger")
		return redirect(url_for("main.dashboard"))
	
	username = request.form.get("username", "").strip()
	confirm_username = request.form.get("confirm_username", "").strip()
	
	# Double confirmation for safety
	if username != confirm_username:
		flash("Username confirmation does not match. User not deleted.", "danger")
		return redirect(url_for("main.dashboard"))
	
	target = User.query.filter_by(username=username).first()
	if not target:
		flash("User not found", "danger")
		return redirect(url_for("main.dashboard"))
	
	# Prevent admin from deleting themselves
	if target.id == current_user.id:
		flash("Cannot delete your own account", "danger")
		return redirect(url_for("main.dashboard"))
	
	# Delete associated transactions first (due to foreign key constraints)
	# Delete transactions where user is the target OR the performer
	WalletTransaction.query.filter(
		(WalletTransaction.user_id == target.id) | 
		(WalletTransaction.performed_by_id == target.id)
	).delete()
	
	# Delete audit logs where user is the actor
	from .models import AuditLog
	AuditLog.query.filter_by(actor_id=target.id).delete()
	
	# Delete user's QR code file if it exists
	if target.qr_filename:
		import os
		qr_path = os.path.join(current_app.root_path, "static", "qr_codes", target.qr_filename)
		if os.path.exists(qr_path):
			try:
				os.remove(qr_path)
			except OSError:
				pass  # Continue even if file deletion fails
	
	# Store user info before deletion for logging and cache invalidation
	user_id = target.id
	user_username = target.username
	
	# Delete the user
	db.session.delete(target)
	db.session.commit()
	
	# Invalidate any cached data for this user
	cache.delete(f"user_data_{user_id}")
	
	flash(f"User {user_username} has been deleted successfully", "success")
	log_event("admin_user_deleted", resource=user_username)
	return redirect(url_for("main.dashboard"))


@main_bp.get("/dashboard")
@login_required
def dashboard():
	print(f"DEBUG: Dashboard accessed by user: {current_user.username}, role: {current_user.role}")
	print(f"DEBUG: User authenticated: {current_user.is_authenticated}")
	
	# Cache user data for 5 minutes to reduce database queries
	cache_key = f"user_data_{current_user.id}"
	user_data = cache.get(cache_key)
	
	if not user_data:
		user_data = {
			'username': current_user.username,
			'balance': current_user.balance,
			'role': current_user.role.value,
			'email': current_user.email
		}
		cache.set(cache_key, user_data, timeout=300)
	
	if current_user.role == Role.ADMIN:
		return render_template("admin_dashboard.html")
	if current_user.role == Role.MANAGER:
		return render_template("manager_dashboard.html")
	# USER
	qr_data_uri = generate_qr_data_uri(current_user)
	return render_template("user_dashboard.html", qr_filename=None, qr_data_uri=qr_data_uri, balance=current_user.balance)


@main_bp.post("/admin/update-balance")
@login_required
def admin_update_balance():
	if current_user.role != Role.ADMIN:
		flash("Unauthorized", "danger")
		return redirect(url_for("main.dashboard"))
	username = request.form.get("username", "").strip()
	delta = int(request.form.get("delta", "0"))
	target = User.query.filter_by(username=username).first()
	if not target:
		flash("User not found", "danger")
		return redirect(url_for("main.dashboard"))
	
	transaction, success, message = change_balance(target, delta, reason="admin_update")
	if success:
		# Invalidate cache for the user whose balance was updated
		cache.delete(f"user_data_{target.id}")
		flash(message, "success")
	else:
		flash(message, "warning")
	return redirect(url_for("main.dashboard"))


@main_bp.post("/manager/update-balance")
@login_required
def manager_update_balance():
	if current_user.role not in [Role.ADMIN, Role.MANAGER]:
		flash("Unauthorized", "danger")
		return redirect(url_for("main.dashboard"))
	username = request.form.get("username", "").strip()
	action = request.form.get("action", "add")
	amount = int(request.form.get("amount", "0"))
	delta = amount if action == "add" else -amount
	target = User.query.filter_by(username=username).first()
	if not target:
		flash("User not found", "danger")
		return redirect(url_for("main.dashboard"))
	
	transaction, success, message = change_balance(target, delta, reason=f"manager_{action}")
	if success:
		# Invalidate cache for the user whose balance was updated
		cache.delete(f"user_data_{target.id}")
		flash(message, "success")
	else:
		flash(message, "warning")
	return redirect(url_for("main.dashboard"))


@main_bp.post("/admin/set-role")
@login_required
def admin_set_role():
	if current_user.role != Role.ADMIN:
		flash("Unauthorized", "danger")
		return redirect(url_for("main.dashboard"))
	username = request.form.get("username", "").strip()
	role_str = request.form.get("role", "user")
	target = User.query.filter_by(username=username).first()
	if not target:
		flash("User not found", "danger")
		return redirect(url_for("main.dashboard"))
	if role_str not in ("user", "manager"):
		flash("Invalid role", "danger")
		return redirect(url_for("main.dashboard"))
	target.role = Role.MANAGER if role_str == "manager" else Role.USER
	db.session.commit()
	flash("Role updated", "success")
	log_event("admin_set_role", resource=target.username, meta=f"role={role_str}")
	return redirect(url_for("main.dashboard"))


@main_bp.post("/admin/bulk-import")
@login_required
def admin_bulk_import():
	if current_user.role != Role.ADMIN:
		flash("Unauthorized", "danger")
		return redirect(url_for("main.dashboard"))
	from openpyxl import load_workbook
	from werkzeug.utils import secure_filename
	from werkzeug.security import generate_password_hash
	from .models import generate_password_from_name

	upload = request.files.get("file")
	if not upload:
		flash("No file uploaded", "danger")
		return redirect(url_for("main.dashboard"))

	wb = load_workbook(upload)
	ws = wb.active
	participants = []

	# Map headers (case-insensitive) so we can find the 'email' column safely
	header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
	# Build a lookup like {'email': idx, 'name': idx}
	header_to_index = {}
	for idx, header_value in enumerate(header_row):
		if header_value is None:
			continue
		normalized = str(header_value).strip().lower()
		header_to_index[normalized] = idx

	# Accept common variants
	email_header_candidates = [
		"email",
		"e-mail",
		"mail",
	]
	name_header_candidates = [
		"name",
		"full name",
		"fullname",
		"participant name",
	]

	# Resolve indices
	email_idx = next((header_to_index[h] for h in email_header_candidates if h in header_to_index), None)
	name_idx = next((header_to_index[h] for h in name_header_candidates if h in header_to_index), None)

	if email_idx is None:
		flash("Could not find 'email' column in the uploaded file.", "danger")
		return redirect(url_for("main.dashboard"))

	for row in ws.iter_rows(min_row=2, values_only=True):
		row_values = list(row)
		email = row_values[email_idx] if email_idx < len(row_values) else None
		name = row_values[name_idx] if (name_idx is not None and name_idx < len(row_values)) else None

		# Skip rows without email
		if not email:
			continue

		# Derive username: prefer cleaned name; fallback to email local-part
		if name:
			username = str(name).strip().replace(" ", "").lower()
		else:
			username = str(email).split("@")[0].strip().replace(" ", "").lower()

		password_plain = generate_password_from_name(username)

		participants.append({
			'name': name,
			'email': email,
			'username': username,
			'password': password_plain,
			'exists': User.query.filter_by(username=username).first() is not None
		})

	return render_template("bulk_import_preview.html", participants=participants)


@main_bp.get("/admin/download-credentials")
@login_required
def admin_download_credentials():
	"""Download user credentials as CSV file"""
	if current_user.role != Role.ADMIN:
		flash("Unauthorized", "danger")
		return redirect(url_for("main.dashboard"))
	
	import csv
	import io
	from datetime import datetime
	
	# Get all users
	users = User.query.all()
	
	# Create CSV content
	output = io.StringIO()
	writer = csv.writer(output)
	
	# Write header
	writer.writerow(['Username', 'Email', 'Role', 'Balance', 'Created At'])
	
	# Write user data
	for user in users:
		writer.writerow([
			user.username,
			user.email or '',
			user.role.value if user.role else '',
			user.balance,
			user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else ''
		])
	
	# Create response
	response = make_response(output.getvalue())
	response.headers['Content-Type'] = 'text/csv'
	response.headers['Content-Disposition'] = f'attachment; filename=user_credentials_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
	
	return response


@main_bp.post("/admin/bulk-import-confirm")
@login_required
def admin_bulk_import_confirm():
	if current_user.role != Role.ADMIN:
		flash("Unauthorized", "danger")
		return redirect(url_for("main.dashboard"))
	from werkzeug.security import generate_password_hash
	from .models import generate_password_from_name

	# Get participants data from form
	participants_list = request.form.getlist("participants")
	if not participants_list:
		flash("No participants data found", "danger")
		return redirect(url_for("main.dashboard"))
	
	import json
	participants = []
	for participant_data in participants_list:
		try:
			# Clean the data if it has extra characters
			clean_data = participant_data.strip()
			if clean_data:
				participant = json.loads(clean_data)
				participants.append(participant)
		except json.JSONDecodeError as e:
			print(f"DEBUG: JSON decode error: {e}")
			print(f"DEBUG: Problematic data: {repr(participant_data)}")
			continue
	
	print(f"DEBUG: Total participants to process: {len(participants)}")
	created = 0
	emails_sent = 0
	
	for participant in participants:
		print(f"DEBUG: Processing participant: {participant}")
		if participant['exists']:
			continue
			
		# Create user
		print(f"DEBUG: Creating user with username: {participant['username']}, email: {participant['email']}")
		user = User(
			username=participant['username'],
			email=participant['email'],
			password_hash=generate_password_hash(participant['password']),
			role=Role.USER,
			balance=0.0
		)
		db.session.add(user)
		db.session.commit()
		ensure_qr_for_user(user)
		
		# Debug logging
		print(f"DEBUG: Created user - username: {participant['username']}, password: {participant['password']}")
		print(f"DEBUG: Password hash: {user.password_hash}")
		print(f"DEBUG: Password check: {check_password_hash(user.password_hash, participant['password'])}")
		
		# Send email asynchronously
		if participant['email']:
			try:
				# Try async email sending first
				from .celery_app import send_email_async
				task = send_email_async.delay(participant['email'], participant['username'], participant['password'])
				print(f"DEBUG: Email queued for {participant['email']} (task: {task.id})")
				emails_sent += 1
			except ImportError:
				# Fallback to synchronous email if Celery not available
				try:
					print(f"DEBUG: Attempting to send email to {participant['email']}")
					send_credentials_email(participant['email'], participant['username'], participant['password'])
					print(f"DEBUG: Email sent successfully to {participant['email']}")
					emails_sent += 1
				except Exception as e:
					print(f"DEBUG: Failed to send email to {participant['email']}: {e}")
					import traceback
					traceback.print_exc()
					flash(f"Failed to send email to {participant['email']}: {str(e)}", "warning")
			except Exception as e:
				print(f"DEBUG: Failed to queue email for {participant['email']}: {e}")
				flash(f"Failed to queue email for {participant['email']}: {str(e)}", "warning")
		
		created += 1

	flash(f"Created {created} users and sent {emails_sent} emails", "success")
	log_event("admin_bulk_import", resource="users", meta=f"created={created}, emails_sent={emails_sent}")
	return redirect(url_for("main.dashboard"))


@main_bp.post("/admin/bulk-import-download")
@login_required
def admin_bulk_import_download():
	"""Download the preview participants list as CSV before sending emails"""
	if current_user.role != Role.ADMIN:
		return jsonify({"error": "Unauthorized"}), 403

	# Get participants data from form (same as confirm)
	participants_list = request.form.getlist("participants")
	if not participants_list:
		flash("No participants data found for download", "danger")
		return redirect(url_for("main.dashboard"))

	import csv
	import io
	import json

	participants = []
	for participant_data in participants_list:
		try:
			clean_data = participant_data.strip()
			if clean_data:
				participant = json.loads(clean_data)
				participants.append(participant)
		except json.JSONDecodeError:
			continue

	# Prepare CSV
	output = io.StringIO()
	writer = csv.writer(output)
	writer.writerow(["Name", "Email", "Username", "Password", "Status"]) 
	for p in participants:
		status = "Already Exists" if p.get("exists") else "Ready"
		writer.writerow([
			p.get("name", ""),
			p.get("email", ""),
			p.get("username", ""),
			p.get("password", ""),
			status,
		])

	response = make_response(output.getvalue())
	response.headers["Content-Type"] = "text/csv"
	response.headers["Content-Disposition"] = "attachment; filename=bulk_import_preview.csv"
	return response

@main_bp.get("/admin/database-monitor")
@login_required
def admin_database_monitor():
	"""Database monitoring dashboard for admin only"""
	if current_user.role != Role.ADMIN:
		flash("Unauthorized", "danger")
		return redirect(url_for("main.dashboard"))
	
	try:
		# Get database connection info
		database_url = os.environ.get("DATABASE_URL")
		if not database_url:
			# Fallback to individual components
			db_host = os.environ.get("DB_HOST", "localhost")
			db_port = os.environ.get("DB_PORT", "5432")
			db_name = os.environ.get("DB_NAME", "token_wallet")
			db_user = os.environ.get("DB_USER", "wallet_user")
			db_password = os.environ.get("DB_PASSWORD", "password")
			database_url = f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}"
		
		# Limits and thresholds from env (defaults for Railway free)
		max_db_size_mb = int(os.environ.get("DB_MAX_SIZE_MB", 1024))  # default 1GB
		warn_threshold = float(os.environ.get("DB_WARN_THRESHOLD", 75))
		critical_threshold = float(os.environ.get("DB_CRITICAL_THRESHOLD", 90))
		
		# Connect to database for monitoring
		conn = psycopg2.connect(database_url)
		cursor = conn.cursor()
		
		# Get database size
		cursor.execute(
			"""
			SELECT pg_size_pretty(pg_database_size(current_database())) as db_size,
			       pg_database_size(current_database()) as db_size_bytes
			"""
		)
		db_size_result = cursor.fetchone()
		db_size_pretty = db_size_result[0] if db_size_result else "Unknown"
		db_size_bytes = db_size_result[1] if db_size_result else 0
		
		# Get table sizes
		cursor.execute(
			"""
			SELECT 
				schemaname,
				tablename,
				pg_size_pretty(pg_total_relation_size(schemaname||'.'||tablename)) as size,
				pg_total_relation_size(schemaname||'.'||tablename) as size_bytes,
				pg_stat_get_tuples_returned(c.oid) as row_count
			FROM pg_tables pt
			JOIN pg_class c ON c.relname = pt.tablename
			WHERE schemaname = 'public'
			ORDER BY pg_total_relation_size(schemaname||'.'||tablename) DESC
			"""
		)
		table_sizes = cursor.fetchall()
		
		# Get connection info
		cursor.execute(
			"""
			SELECT 
				count(*) as total_connections,
				count(*) FILTER (WHERE state = 'active') as active_connections,
				count(*) FILTER (WHERE state = 'idle') as idle_connections
			FROM pg_stat_activity 
			WHERE datname = current_database()
			"""
		)
		connection_info = cursor.fetchone()
		
		# Get user count
		user_count = User.query.count()
		
		# Get recent activity (last 24 hours)
		cursor.execute(
			"""
			SELECT 
				date_trunc('hour', created_at) as hour,
				count(*) as transactions
			FROM transaction 
			WHERE created_at >= NOW() - INTERVAL '24 hours'
			GROUP BY date_trunc('hour', created_at)
			ORDER BY hour DESC
			LIMIT 24
			"""
		)
		recent_activity = cursor.fetchall()
		
		# Calculate usage percentage with env-based max size
		max_db_size_bytes = max_db_size_mb * 1024 * 1024
		usage_percentage = (db_size_bytes / max_db_size_bytes) * 100 if db_size_bytes > 0 else 0
		
		# Determine status using env thresholds
		if usage_percentage >= critical_threshold:
			status = "CRITICAL"
			status_color = "danger"
		elif usage_percentage >= warn_threshold:
			status = "WARNING"
			status_color = "warning"
		elif usage_percentage >= 50:
			status = "MODERATE"
			status_color = "info"
		else:
			status = "HEALTHY"
			status_color = "success"
		
		cursor.close()
		conn.close()
		
		return render_template(
			"admin_database_monitor.html",
			db_size=db_size_pretty,
			db_size_bytes=db_size_bytes,
			usage_percentage=usage_percentage,
			max_db_size_mb=max_db_size_mb,
			status=status,
			status_color=status_color,
			table_sizes=table_sizes,
			connection_info=connection_info,
			user_count=user_count,
			recent_activity=recent_activity
		)
		
	except Exception as e:
		flash(f"Database monitoring error: {str(e)}", "danger")
		return redirect(url_for("main.dashboard"))


@main_bp.get("/admin/database-monitor/api")
@login_required
def admin_database_monitor_api():
	"""API endpoint for real-time database monitoring"""
	if current_user.role != Role.ADMIN:
		return jsonify({"error": "Unauthorized"}), 403
	
	try:
		# Get database connection info
		database_url = os.environ.get("DATABASE_URL")
		if not database_url:
			db_host = os.environ.get("DB_HOST", "localhost")
			db_port = os.environ.get("DB_PORT", "5432")
			db_name = os.environ.get("DB_NAME", "token_wallet")
			db_user = os.environ.get("DB_USER", "wallet_user")
			db_password = os.environ.get("DB_PASSWORD", "password")
			database_url = f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}"
		
		# Limits and thresholds from env
		max_db_size_mb = int(os.environ.get("DB_MAX_SIZE_MB", 1024))
		warn_threshold = float(os.environ.get("DB_WARN_THRESHOLD", 75))
		critical_threshold = float(os.environ.get("DB_CRITICAL_THRESHOLD", 90))
		
		conn = psycopg2.connect(database_url)
		cursor = conn.cursor()
		
		# Get current database size
		cursor.execute("SELECT pg_database_size(current_database())")
		db_size_bytes = cursor.fetchone()[0]
		
		# Get connection count
		cursor.execute(
			"""
			SELECT count(*) FROM pg_stat_activity 
			WHERE datname = current_database()
			"""
		)
		connection_count = cursor.fetchone()[0]
		
		# Get user count
		user_count = User.query.count()
		
		# Calculate usage percentage
		usage_percentage = (db_size_bytes / (max_db_size_mb * 1024 * 1024)) * 100
		
		cursor.close()
		conn.close()
		
		return jsonify({
			"db_size_bytes": db_size_bytes,
			"db_size_pretty": f"{db_size_bytes / (1024*1024):.2f} MB",
			"usage_percentage": round(usage_percentage, 2),
			"connection_count": connection_count,
			"user_count": user_count,
			"status": "CRITICAL" if usage_percentage >= critical_threshold else "WARNING" if usage_percentage >= warn_threshold else "HEALTHY",
			"timestamp": time.time()
		})
		
	except Exception as e:
		return jsonify({"error": str(e)}), 500


